from dataclasses import asdict, dataclass
from itertools import product
from pathlib import Path

import esptest.common.compat_typing as t

from ..logger import get_logger

VarType: t.TypeAlias = t.Union[int, float, str]
logger = get_logger('iperf-util')

try:
    from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
    from openpyxl.styles import Font  # pyright: ignore[reportMissingModuleSource]

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


_PHY_PREFIXES = ('11ax', '11ac', '11n', '11b', '11g', 'lr')
_PHY_MODE_ORDER = ('11b', '11g', '11n', '11ax', '11ac', 'lr', 'other')
_RATE_ORDER = {
    '11b': ['1M', '2M', '5.5M', '11M', 'auto'],
    '11g': ['6M', '9M', '12M', '18M', '24M', '36M', '48M', '54M', 'auto'],
    '11n': ['MCS0', 'MCS1', 'MCS2', 'MCS3', 'MCS4', 'MCS5', 'MCS6', 'MCS7', 'MCS7s', 'auto'],
    '11ax': ['MCS0', 'MCS1', 'MCS2', 'MCS3', 'MCS4', 'MCS5', 'MCS6', 'MCS7', 'MCS8', 'MCS9', 'auto'],
    '11ac': [
        'VHT_MCS0',
        'VHT_MCS1',
        'VHT_MCS2',
        'VHT_MCS3',
        'VHT_MCS4',
        'VHT_MCS5',
        'VHT_MCS6',
        'VHT_MCS7',
        'VHT_MCS8',
        'VHT_MCS9',
        'auto',
    ],
    'lr': ['LORA_250K', 'LORA_500K', 'auto'],
}
_THROUGHPUT_INCREASE_TOLERANCE = 0.05


@dataclass
class FixRateReportOptions:
    """Formatting options shared by the fix-rate report generators."""

    init_rssi: int = 0
    ap_name: str = ''
    target: str = ''
    throughput_value: str = 'max'


def parse_rate_label(label: str) -> tuple:
    """Parse config/rate labels like ``perf_rate_11n_MCS0`` or ``11n_MCS0``."""
    label = str(label or '')
    if '_rate_' in label:
        label = label.split('_rate_', 1)[1]
    lower = label.lower()
    for phy in _PHY_PREFIXES:
        prefix = phy + '_'
        if lower.startswith(prefix):
            return phy, label[len(prefix) :]
        if lower == phy:
            return phy, label
    return 'other', label


def _rate_sort_key(phy_mode: str, rate_name: str) -> tuple:
    order = _RATE_ORDER.get(phy_mode, [])
    try:
        return (0, order.index(rate_name))
    except ValueError:
        return (1, rate_name)


@dataclass
class IperfResult:
    """One point iperf result, including type, att, rssi, max, avg, min, heap, etc."""

    avg: float
    max: float = -1  # Can be ignored
    min: float = -1  # Can be ignored
    throughput_list: t.Optional[t.List[float]] = None
    unit: str = 'Mbits/sec'
    min_heap: int = 0
    bandwidth: int = 0  # 20/40
    channel: int = 0
    rssi: float = -128
    type: str = 'iperf'  # tcp_tx, tcp_rx, udp_tx, udp_rx
    target: str = ''  # esp32, esp32s2, etc.
    errors: t.Optional[t.List[str]] = None
    # Advanced
    att: int = 0
    config_name: str = 'unknown'
    ap_name: str = 'unknown'
    version: str = 'unknown'

    def to_dict(self, with_keys: t.Optional[t.List[str]] = None) -> t.Dict[str, float]:
        """_summary_

        Args:
            with_keys (Optional[List[str]], optional): dict keys. default [avg,max,min,min_heap,rssi].
        """
        if not with_keys:
            with_keys = ['avg', 'max', 'min', 'min_heap', 'rssi']

        d = {}
        for k, v in asdict(self).items():
            if k not in with_keys:
                continue
            if not isinstance(v, (int, float)):
                logger.error(f'Variable of {k} must be a number, got {v}')
                raise ValueError(f'Variable of {k} must be a number, got {v}')
            d[k] = v
        return d


class IperfResultsRecord:
    """record, analysis iperf test results for different configs"""

    def __init__(self) -> None:
        self._results: t.List[IperfResult] = []
        self._aps: t.Set[str] = set()
        self._targets: t.Set[str] = set()
        self._types: t.Set[str] = set()

    def append_result(self, result: IperfResult) -> None:
        self._results.append(result)
        self._aps.add(result.ap_name)
        self._targets.add(result.target)
        self._types.add(result.type)

    def part(self, filter_fn: t.Callable[[IperfResult], bool]) -> 't.Self':
        new_record = self.__class__()
        for result in self._results:
            if filter_fn(result):
                continue
            new_record.append_result(result)
        return new_record

    def _dict_by_key(
        self,
        key: str,
        filter_fn: t.Optional[t.Callable[[IperfResult], bool]] = None,
        reverse: bool = False,
    ) -> t.Dict[VarType, t.List[IperfResult]]:
        if not self._results:
            raise ValueError('No iperf test results recorded.')
        key_list = list(sorted({getattr(r, key) for r in self._results}, reverse=reverse))
        if len(key_list) <= 1:
            logger.info(f'Did not find different {key} in iperf test results.')
        d_key: t.Dict[VarType, t.List[IperfResult]] = {k: [] for k in key_list}
        for res in self._results:
            if filter_fn and not filter_fn(res):
                continue
            d_key[getattr(res, key)].append(res)
        return d_key

    def dict_by_att(
        self, filter_fn: t.Optional[t.Callable[[IperfResult], bool]] = None
    ) -> t.Dict[VarType, t.List[IperfResult]]:
        return self._dict_by_key('att', filter_fn)

    def dict_by_ap(
        self, filter_fn: t.Optional[t.Callable[[IperfResult], bool]] = None
    ) -> t.Dict[VarType, t.List[IperfResult]]:
        return self._dict_by_key('ap_name', filter_fn)

    def _format_label_str(self, base_label: str, ap_name: str = '', target: str = '') -> str:
        assert base_label
        labels = []
        if len(self._aps) > 1:
            labels.append(ap_name)
        if len(self._targets) > 1:
            labels.append(target)
        labels.append(base_label)
        label_str = '_'.join(labels)
        return label_str

    @staticmethod
    def _get_matched_result(
        from_results: t.Iterable[IperfResult], filter_fn: t.Callable[[IperfResult], bool]
    ) -> t.Optional[IperfResult]:
        """Get first matched result by given condition"""
        for result in from_results:
            if filter_fn(result):
                return result
        return None

    def draw_rssi_vs_att_chart(
        self,
        file_name: str,
        title: str = 'RSSI vs ATT',
    ) -> None:
        # Needs extra packages to draw the chart
        from .line_chart import draw_line_chart_basic

        raw_data = self.dict_by_att()

        x_data: t.List[int] = []
        y_data: t.List[t.Dict[str, t.Optional[float]]] = []
        for att, results in raw_data.items():
            assert isinstance(att, int)
            x_data.append(att)
            _data: t.Dict[str, t.Optional[float]] = {}
            for ap, target in product(self._aps, self._targets):
                # pylint: disable=cell-var-from-loop
                label = self._format_label_str('rssi', ap, target)
                result = self._get_matched_result(results, lambda res: res.ap_name == ap and res.target == target)
                if result:
                    _data[label] = result.rssi
                else:
                    _data[label] = None
            y_data.append(_data)
        draw_line_chart_basic(file_name, title, y_data, x_data, x_label='att', y_label='rssi')

    def draw_rate_vs_rssi_chart(
        self,
        file_name: str,
        title: str = 'Rate vs RSSI',
        throughput_type: str = 'max',
    ) -> None:
        # Needs extra packages to draw the chart
        from .line_chart import draw_line_chart_basic

        # draw rssi chart from high rssi to low rssi
        raw_data = self._dict_by_key('rssi', reverse=True)

        x_data: t.List[float] = []
        y_data: t.List[t.Dict[str, t.Optional[float]]] = []
        for rssi, results in raw_data.items():
            assert isinstance(rssi, (int, float))
            x_data.append(-rssi)  # left value is higher rssi
            _data: t.Dict[str, t.Optional[float]] = {}
            for ap, target, typ in product(self._aps, self._targets, self._types):
                # pylint: disable=cell-var-from-loop
                label = self._format_label_str(typ, ap, target)
                result = self._get_matched_result(
                    results, filter_fn=lambda res: res.ap_name == ap and res.target == target and res.type == typ
                )
                if result:
                    _data[label] = result.avg if throughput_type == 'avg' else result.max
                else:
                    _data[label] = None
            y_data.append(_data)
        draw_line_chart_basic(file_name, title, y_data, x_data, x_label='rssi (-)', y_label='rate')

    @staticmethod
    def _throughput_value(result: IperfResult, throughput_type: str) -> float:
        if throughput_type == 'avg':
            return result.avg
        if throughput_type == 'min':
            return result.min
        return result.max

    @staticmethod
    def _format_throughput(value: t.Optional[float], highlight: bool = False) -> str:
        if value is None:
            return ''
        text = f'{value:.2f} Mbps'
        return f'<font color="red">{text}</font>' if highlight else text

    @staticmethod
    def _is_unexpected_increase(previous_throughput: t.Optional[float], throughput: t.Optional[float]) -> bool:
        if previous_throughput is None or throughput is None:
            return False
        return throughput > previous_throughput * (1 + _THROUGHPUT_INCREASE_TOLERANCE)

    @staticmethod
    def _is_auto_rate(rate_name: str) -> bool:
        return str(rate_name or '').lower() == 'auto'

    @staticmethod
    def _is_auto_below_fixed_max(fixed_max_throughput: t.Optional[float], throughput: t.Optional[float]) -> bool:
        if fixed_max_throughput is None or throughput is None:
            return False
        return throughput < fixed_max_throughput * (1 - _THROUGHPUT_INCREASE_TOLERANCE)

    def _fix_rate_groups(
        self,
        throughput_type: str,
        ap_name: str = '',
        target: str = '',
    ) -> t.Dict[str, t.List[tuple]]:
        groups: t.Dict[str, t.List[tuple]] = {}
        for result in self._results:
            if result.type != throughput_type:
                continue
            if ap_name and result.ap_name != ap_name:
                continue
            if target and result.target != target:
                continue
            phy, rate_name = parse_rate_label(result.config_name)
            groups.setdefault(phy, []).append((rate_name, result))
        for phy, rows in groups.items():
            # pylint: disable=cell-var-from-loop
            rows.sort(key=lambda item: _rate_sort_key(phy, item[0]))
        return groups

    @staticmethod
    def _result_for_att(results: t.Iterable[IperfResult], att: int) -> t.Optional[IperfResult]:
        for result in results:
            if result.att == att:
                return result
        return None

    @staticmethod
    def _format_rssi_values(results: t.Iterable[IperfResult], att: int) -> str:
        values = []
        for result in results:
            if result.att == att and result.rssi not in values:
                values.append(result.rssi)
        return '/'.join(str(v) for v in values)

    def _fix_rate_rows(
        self,
        rows: t.List[tuple],
        atts: t.List[int],
        throughput_value: str,
    ) -> t.List[t.Tuple[str, t.List[t.Tuple[t.Optional[float], bool]]]]:
        """Build ``(rate_name, [(throughput, highlight), ...])`` for each unique rate in display order."""
        fixed_max_by_att: t.Dict[int, t.Optional[float]] = {att: None for att in atts}
        rate_names: t.List[str] = []
        for rate_name, _result in rows:
            if rate_name not in rate_names:
                rate_names.append(rate_name)
        result_rows: t.List[t.Tuple[str, t.List[t.Tuple[t.Optional[float], bool]]]] = []
        for rate_name in rate_names:
            same_rate_results = [r for rn, r in rows if rn == rate_name]
            is_auto_rate = self._is_auto_rate(rate_name)
            previous_throughput: t.Optional[float] = None
            cells: t.List[t.Tuple[t.Optional[float], bool]] = []
            for att in atts:
                result = self._result_for_att(same_rate_results, att)
                current = self._throughput_value(result, throughput_value) if result is not None else None
                highlight = self._is_unexpected_increase(previous_throughput, current) or (
                    is_auto_rate and self._is_auto_below_fixed_max(fixed_max_by_att.get(att), current)
                )
                cells.append((current, highlight))
                if current is None:
                    continue
                previous_throughput = current
                if not is_auto_rate:
                    fixed_max = fixed_max_by_att.get(att)
                    if fixed_max is None or current > fixed_max:
                        fixed_max_by_att[att] = current
            result_rows.append((rate_name, cells))
        return result_rows

    def _fix_rate_markdown_phy(self, phy: str, rows: t.List[tuple], opts: 'FixRateReportOptions') -> str:
        atts = sorted({result.att for _rate_name, result in rows})
        data = f'\r\n### {phy}:\r\n\r\n'
        data += '| ATT |' + ''.join(f' {att} |' for att in atts) + '\r\n'
        data += '|------|' + ''.join('------------------|' for _ in atts) + '\r\n'
        data += '| Theoretical RSSI |' + ''.join(f' {opts.init_rssi - att} |' for att in atts) + '\r\n'
        data += (
            '| Scanned RSSI |'
            + ''.join(f' {self._format_rssi_values((r for _rn, r in rows), att)} |' for att in atts)
            + '\r\n'
        )
        for rate_name, cells in self._fix_rate_rows(rows, atts, opts.throughput_value):
            row = f'| {rate_name} |'
            for throughput, highlight in cells:
                row += f' {self._format_throughput(throughput, highlight)} |'
            data += row + '\r\n'
        return data

    def _fix_rate_markdown_section(self, tp_type: str, opts: 'FixRateReportOptions') -> str:
        data = f'\r\n## {tp_type}\r\n'
        groups = self._fix_rate_groups(tp_type, opts.ap_name, opts.target)
        has_data = False
        for phy in _PHY_MODE_ORDER:
            if phy not in groups:
                continue
            has_data = True
            data += self._fix_rate_markdown_phy(phy, groups[phy], opts)
        if not has_data:
            data += '\r\n(no data)\r\n'
        return data

    def generate_fix_rate_raw_data_markdown(
        self,
        throughput_types: t.Optional[t.List[str]] = None,
        options: t.Optional['FixRateReportOptions'] = None,
    ) -> str:
        """Generate fix-rate raw-data markdown aligned with QACT report format."""
        opts = options or FixRateReportOptions()
        if throughput_types is None:
            throughput_types = sorted(self._types)
        data = '# Fix Rate Raw Data\r\n'
        if opts.ap_name:
            data += f'\r\nAP: {opts.ap_name}\r\n'
        if opts.target:
            data += f'\r\nTarget: {opts.target}\r\n'
        data += f'\r\ninit_rssi: {opts.init_rssi}\r\n'
        for tp_type in throughput_types:
            data += self._fix_rate_markdown_section(tp_type, opts)
        return data

    def save_fix_rate_raw_data_markdown(
        self,
        file_name: str,
        throughput_types: t.Optional[t.List[str]] = None,
        options: t.Optional['FixRateReportOptions'] = None,
    ) -> None:
        data = self.generate_fix_rate_raw_data_markdown(throughput_types, options)
        path = Path(file_name)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(data, encoding='utf-8')

    def _write_fix_rate_excel_sheet(self, wb: 'Workbook', tp_type: str, opts: 'FixRateReportOptions') -> None:
        ws = wb.create_sheet(title=tp_type[:31])
        groups = self._fix_rate_groups(tp_type, opts.ap_name, opts.target)
        row = 1
        for phy in _PHY_MODE_ORDER:
            if phy not in groups:
                continue
            rows = groups[phy]
            atts = sorted({result.att for _rate_name, result in rows})
            ws.cell(row=row, column=1, value=phy)
            row += 1
            ws.cell(row=row, column=1, value='ATT')
            for col, att in enumerate(atts, start=2):
                ws.cell(row=row, column=col, value=att)
            row += 1
            ws.cell(row=row, column=1, value='Theoretical RSSI')
            for col, att in enumerate(atts, start=2):
                ws.cell(row=row, column=col, value=opts.init_rssi - att)
            row += 1
            ws.cell(row=row, column=1, value='Scanned RSSI')
            for col, att in enumerate(atts, start=2):
                ws.cell(row=row, column=col, value=self._format_rssi_values((r for _rn, r in rows), att))
            row += 1
            for rate_name, cells in self._fix_rate_rows(rows, atts, opts.throughput_value):
                ws.cell(row=row, column=1, value=rate_name)
                for col, (throughput, highlight) in enumerate(cells, start=2):
                    cell = ws.cell(row=row, column=col, value=self._format_throughput(throughput, highlight=False))
                    if highlight:
                        cell.font = Font(color='FF0000')
                row += 1
            row += 1

    def save_fix_rate_raw_data_excel(
        self,
        file_name: str,
        throughput_types: t.Optional[t.List[str]] = None,
        options: t.Optional['FixRateReportOptions'] = None,
    ) -> None:
        """Save fix-rate raw data as Excel. Requires optional openpyxl."""
        if not HAS_OPENPYXL:
            raise ImportError('openpyxl is required to save fix-rate Excel reports')
        opts = options or FixRateReportOptions()
        if throughput_types is None:
            throughput_types = sorted(self._types)
        wb = Workbook()
        wb.remove(wb.active)
        for tp_type in throughput_types:
            self._write_fix_rate_excel_sheet(wb, tp_type, opts)
        path = Path(file_name)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
